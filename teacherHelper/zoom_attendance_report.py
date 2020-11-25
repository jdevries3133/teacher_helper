from copy import copy
import datetime
import logging
import string
import json
import re

from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

try:
    from .manual_zoom_matches import MANUAL_FIXES
except ImportError:
    def MANUAL_FIXES(name):
        """
        For students with silly names the algorithm cannot recognize, write
        this function in the file above to make individual corrections.
        ./manual_zoom_matches.py is in the .gitignore because it will
        contain student names.
        """
        return name


from .helper import Helper

logger = logging.getLogger(__name__)


class ZoomReportFormatIncorrect(Exception):
    """
    Raised by Meeting when the user apparently didn't check the boxes they
    needed to check when generating the report.
    """


class HelperConsumer:
    helper = Helper.read_cache()


class Meeting(HelperConsumer):
    """
    Single zoom meeting instance. Takes a string of csv data, and parses out
    the following attributes:

    Attrs:
        - topic         str
        - datetime      datetime.datetime
        - attendees     list[teacherHelper.helperStudent]

    Operator overloading on these classes is inconsistent:

        - if self == other: they have the same datetime
        - self > other: self has more attendees.
    """

    def __init__(self, csv_string: str, known_matches=None):
        super().__init__()
        self.csv_string = csv_string

        # known_matches is a cache layer, passed down from MeetingSet
        self.known_matches = known_matches if known_matches else {}
        self.SEARCH_CONFIDENCE_THRESHOLD = 90

        self.average_len_attendance = None  # int
        self.attendees = []                 # list[helper.Student]
        self.unidentifiable = []            # list[str]
        self.datetime = None                # datetime.datetime
        self.topic = None                   # str

        # may not be assigned; only used for matching
        self.grade_level = None             # int
        self.homeroom = None                # str
        self.rows = []                      # raw rows  list[str] (does not persist across MeetingSet serialization)

        # first pass matches that will be skipped on the second pass
        self._matched = []

    def __repr__(self):
        outstr = (
            f'<self.helper.zoom_attendance_report.Meeting; {self.topic} at '
            f'{self.datetime.isoformat()}>'
        )
        return outstr

    def __str__(self):
        return f'{self.datetime.isoformat()}; {self.topic}'

    def __eq__(self, other):
        """
        Meetings are equal if they happened at the same datetime and have the
        same topic. This can be simplified to str(self) == str(other), because
        Meeting.__str__ outputs both these values.
        """
        # can only compare meeting to other Meetings
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        return str(self) == str(other)

    def __gt__(self, other):
        """
        Based on length of attendees.
        """
        # assert type(self) == type(other) or raise ValueError
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        return len(self.attendees) > len(other.attendees)

    def __len__(self):
        return len(self.attendees)

    def read_report(self):
        """
        Reads, and parses zoom report string.

        assigns attributes:
        self.topic
        self.grade_level
        self.datetime
        """
        self._validate_zoom_report_format()
        self._parse_csv()

    def _validate_zoom_report_format(self):
        """
        Validate that the user checked the boxes they needed to to check while
        generating the CSV
        """
        self.rows = []
        for line in self.csv_string.split('\n')[:-1]:
            line = line.strip()
            self.rows.append(line.split(','))
        # raise an exception if row 2 is not blank.
        # tolerate the case of row[2] = ['', '', '', ...]
        exception_message = (
            'Zoom report must contain meeting information. The report for '
            f'{self.topic} at {self.datetime} does not appear to contain '
            'meeting information.'
        )
        if self.rows[2]:
            row_set = set(self.rows[2])  # this line should be blank
            if len(row_set) <= 1:
                if row_set.pop():
                    raise ZoomReportFormatIncorrect(exception_message)
            else:
                raise ZoomReportFormatIncorrect(exception_message)

    def _parse_csv(self):
        """
        Assuming that it's structure has been validated.
        """
        self._parse_csv_header()
        self._parse_csv_body()

    def _parse_csv_header(self):
        self.topic = self.rows[1][1]
        time_str = self.rows[1][2]
        (
            month,
            day,
            year,
            hour,
            minute,
            *rest  # pylint: disable=unused-variable
        ) = [int(i) for i in re.split(r'/| |:', time_str) if i.isdigit()]
        if 'pm' in time_str.lower():
            hour += 12
        self.datetime = datetime.datetime(year, month, day, hour, minute)
        logger.info(f'{"*" * 30} PARSING {self.topic} at {self.datetime.date()} {"*" * 30}')

    def _parse_csv_body(self):
        """
        Make two passes over the data. Once for the easy-to-match students,
        and again using data gained from the easy to match students to narrow
        field of search for the hard to match ones.
        """
        self._csv_parse_first_pass()
        self._csv_parse_second_pass()

    def _csv_parse_first_pass(self):
        """
        First patch fetching only high confidence matches and determining
        subgroups within which we can search.
        """
        logger.debug('*** First matching pass ***')
        grade_levels_within = set()
        homerooms_within = set()
        for row in self.rows[4:]:

            # try fetching from cache layer
            if not (st := self.known_matches.get(row[0])):

                # if that doesn't work, use high reliability search from
                # ./helper.Helper
                st = self.helper.find_nearest_match(
                    row[0],
                    auto_yes=True,
                    threshold=self.SEARCH_CONFIDENCE_THRESHOLD)

                # if that doesn't work, try again on the next pass
                if not st:
                    continue

                self.known_matches[row[0]] = st  # put new match into cache

            else:
                logger.debug(f'Global cache hit for {st.name}')

            logger.debug(f'FIRST PASS MATCH {row[0]} == {st.name}')

            st.zoom_attendance_report.setdefault(
                self.__str__(),
                int(row[2])  # duration of attendance
            )
            self._matched.append(row[0])
            grade_levels_within.add(st.grade_level)
            homerooms_within.add(st.homeroom)

        logger.debug(
            f'After the first matching pass, {len(self._matched)} students have been '
            'matched.\n*** Second matching pass ***'
        )

        # we may have determined grade level or homerooms, thus narrowing the
        # search
        if len(homerooms_within) <= 1:
            self.homeroom = homerooms_within.pop()
        elif len(grade_levels_within) <= 1:  # elif because homeroom comparison is preferable anyway
            self.grade_level = grade_levels_within.pop()

        logger.debug(f'Presumed grade_level is\t{self.grade_level}')
        logger.debug(f'Presumed homeroom is\t{self.homeroom}')

    def _csv_parse_second_pass(self):
        """
        Make a second pass over the data, matching within subgroup.
        """
        for row in self.rows[4:]:

            # skip those we've already matched
            if row[0] in self._matched:
                continue

            logger.debug(f'Attempting to match {row[0]} on the second pass')

            # try to match upside down, forwards and backwards
            st = self.match_student(row[0])

            if not st:
                logger.debug(f'No match for {row[0]}')
                self.unidentifiable.append(row[0])
                continue

            # It's a match!
            logger.debug(f'SECOND PASS MATCH {row[0]} == {st.name}')
            st.zoom_attendance_report.setdefault(
                self.__str__(),
                int(row[2])  # duration of attendance
            )
            self.attendees.append(st)

    def match_student(self, name):
        """
        Wrapper method that ties together the self.helper class's find_student
        method, the try_matching_student... method in this class, and some
        custom matches for students with really weird zoom names. Here is also
        where the student's raw zoom name is sanitized (punctuation removed).
        """
        # reference manual fixes, an optional import
        name = MANUAL_FIXES(name)

        # Rough cleaning
        # some use dot to delimit first / last name
        name.replace('.', ' ')
        # remove punctuation
        for char in string.punctuation:
            name.replace(char, '')
        # popular emoticon character
        name.replace('ω', '')
        try:
            # use helper search, which can raise warnings
            st = self.helper.find_nearest_match(
                name,
                auto_yes=True,
                threshold=self.SEARCH_CONFIDENCE_THRESHOLD
            )
            if st:
                return st
        except Warning:
            pass
        # search whole name wthin subgroup
        st = self.try_matching_student_within_subgroup(name)
        if st:
            return st
        for part in [i for i in re.split(r'[ |.]', name) if i]:
            # search each word in name within subgroup
            st = self.try_matching_student_within_subgroup(
                part,
                name_part='first_name'
            )
            if st:
                return st
            # if that didn't work, try again assuming that this part is a last
            # name
            st = self.try_matching_student_within_subgroup(
                part,
                name_part='last_name'
            )
            return st

    # pylint: disable=inconsistent-return-statements
    def try_matching_student_within_subgroup(self, student_name, name_part='name'):
        """
        Unlike in the self.helper student matching function, this class is aware of
        the grade level of the student it is trying to match. If the self.helper
        method returns None, this fallback method tries to identify the
        student through process of elimination within their own grade. This
        helps match more students.

        name_part is the part of the student_name being provided, and the
        part that we will search against for the other names. Acceptable values
        are:

            - 'first_name'
            - 'last_name'
            - 'name' (full name; first and last)
        """

        # ok this is super confusing so here goes... there's definitely gonna
        # be more explaining here than actual code....

        # first, recognize that self.homeroom and self.grade_level are
        # initialized as None.

        # now, self._parse_csv_body() iterates over the rows twice. On the
        # first pass, this function is NOT called. Instead,
        # teacherHelper.helper.Helper.find_nearest_match is used ONLY. The
        # reason for this is that it provides more reliable results, but
        # also has a higher likelihood of failure.

        # on the second pass through the data, this func is called in a last
        # ditch attempt to find a match for the student. See, during the first
        # pass, we created a set of the homerooms and grade levels that the
        # easy-to-match students were part of. IF this set is only one item
        # long after the easier half of the group is matched, we assume
        # that the rest of the group is also part of that subgroup.

        # With that assumption made, self.homeroom and self.grade_level are set
        # to a string and integer, respectively. Now, we can use those values
        # to filter all of the students in self.helper.students() and search
        # within a smaller subgroup instad. THAT is what happens here.

        # The "magic" in this function is that we may be seraching within the
        # homeroom, or we may be searching within the whole grade level – we
        # don't know yet when the function is called. This first block of code
        # determines which filter is available to us, and then produces a
        # queryset of student names from that subgroup which we can fuzzily
        # match against.

        if self.homeroom:  # preferentially search within homeroom
            compare_attr = 'homeroom'
        elif self.grade_level:  # fallback on grade level
            compare_attr = 'grade_level'
        else:
            return  # no subgroup to compare within

        logger.debug('Searching for subgroup match')
        logger.debug(f'Name: {student_name}')
        logger.debug(f'Name Part: {name_part}')
        logger.debug(f'Subgroup: {compare_attr}')

        if (
            st := (
                self._check_scoped_cache(
                    getattr(self, compare_attr), student_name)
            )
        ):
            logger.debug(f'Subgroup-scoped cache hit for {st.name}')
            return st
        qs = [
            getattr(s, name_part) for s in self.helper.students.values()
            if getattr(s, compare_attr) == getattr(self, compare_attr)
        ]
        name_match = process.extract(
            student_name,
            qs,
            limit=2
        )

        # Do not return low confidence match
        if not name_match[0][1] > self.SEARCH_CONFIDENCE_THRESHOLD:
            return

        # Proceed with potential match
        if name_match[1][0] == name_match[0][0]:
            logger.debug(
                f'Cannot proceed with {student_name}. More than one '
                f'student in the {self.grade_level}th grade has the first '
                f'name {name_match[0][0]}.'
            )

        name = name_match[0][0]

        if name == 'name':  # shortcut for full name matches
            if (st := self.helper.students[name]):
                return st

        # we know the name is unique in the subgroup, so find the corresponding
        # Student object in the subgroup and return it.
        for st in self.helper.students.values():

            # skip if st is not in the subgroup
            if not (
                getattr(st, compare_attr) == getattr(self, compare_attr)
            ):
                continue

            # break when we find the right student
            if name in st.name:
                break
        logger.debug(
            f'SUBGROUP MATCH {name} matches with {st.name} within '
            + compare_attr
        )

        # insert into cache, scoped to this subgroup
        self.known_matches[getattr(self, compare_attr)].setdefault(
            student_name,
            st
        )
        return st

    def _check_scoped_cache(self, cache_key, student_name):
        """
        cache_key will be the name of the subgroup. Returns Student or
        None.
        """
        scoped_cache = self.known_matches.get(cache_key)
        if not scoped_cache:
            self.known_matches[cache_key] = {}
            return
        return scoped_cache.get(student_name)




class MeetingSet(HelperConsumer):
    """

    # General Usage

    This class dynamically groups meetings by attendees. For example,
    if you use the same meeting topic (i.e. "Health") to meet with different
    groups of homerooms throughout the week, this class will observe the union
    of the set of attendees at each meeting instance, and if there is a
    significant intersection, those meetings are grouped.

    # Resultant Data Structure

    Meeting groupings will be accessible as a nested list (self.groups).
    Each item in the list will itself be a chronologically sorted list of
    Meeting instances. A flattened list of all meetings is also available at
    self.meetings for convenience.

    The presumption is that there is no reliable way to know the full name of
    a meeting. For example, the meeting topic might be, "Health," but whoose
    homeroom is it? There's no way to know from here. Hence, this class simply
    takes a large pool of meetings and groups them. Labeling can be done
    elsewhere.

    The topics of each meeting are saved as an attribute on the
    meeting; "meeting.topic". It is possible for multiple meeting instances
    with the same topic to be split into different groups if the participants
    are different enough. This is expected such as in the example case detailed
    above. However, meetings with different topics will never be grouped
    together.

    If, for whatever reason, a teacher has meet with the same class under
    different zoom meet topics, that case could probably be dealt with in a
    more flexible subclass which modifies the "match_meeting_with_group_by_union"
    method.

    # What is self.TOTAL_TO_UNION_RATIO_ADJUSTMENT?

    In a perfect world with perfect attendance, the union between like groups
    will be half of the total number of attendees when comparing two meeting
    instances. In a garbage world where barely anyone consistently attends,
    we would at least expect the union between like groups to be less than
    the total.

    By knocking the total down by 15%, we can find a middle ground. It
    basically allows us to say, "if most of the students from this meeting are
    the same as the ones from the past meeting, then it's a match."
    """


    def __init__(self, csv_strings: list, group_map=None, trust_topics=False):
        super().__init__()
        self.csv_strings = csv_strings
        self.groups = []
        self.meetings = []  # all meetings in a flattened list
        self.known_matches = {}  # match cache
        self.TOTAL_TO_UNION_RATIO_ADJUSTMENT = 0.9
        self.is_processed = False

        # init dict on student objects
        for st in self.helper.students.values():
            st.__dict__.setdefault('zoom_attendance_report', {})

    def process(self):
        """
        Generator that produces data structure. Yields a meeting immediately
        after it has been parsed.
        """
        for csv_string in self.csv_strings:
            meeting = Meeting(csv_string, known_matches=self.known_matches)
            meeting.read_report()
            self.known_matches = {
                **meeting.known_matches,
                **self.known_matches
            }
            self.meetings.append(meeting)
            match = self.match_meeting_with_group_by_union(meeting)
            if match:
                match.append(meeting)
            else:
                self.groups.append([meeting])
            yield meeting  # report progress up the callstack.

        self.is_processed = True

    def match_meeting_with_group_by_union(self, meeting: Meeting):
        """
        Given a list of students, calculate the union between that list, and
        all the other lists of students in meetings previously provided.

        Return the group whose union against the provided list is less than the
        length of the lists combined, indicating that these are two instances
        of the same group of students meeting.
        """
        for group in self.groups:
            # select closest_meeting: the past meeting whose len() is closest
            # to the current meeting.
            closest_meeting = None
            smallest_difference = 100  # initialize with random large number
            for prev_meeting in group:
                diff = abs(len(prev_meeting) - len(meeting))
                if diff < smallest_difference:
                    smallest_difference = diff
                    closest_meeting = prev_meeting

            # perform set union comparison
            cm__attendees = {s.name for s in closest_meeting.attendees}
            m__attendees = {s.name for s in meeting.attendees}
            union = len(cm__attendees | m__attendees)
            total = len(cm__attendees) + len(m__attendees)
            total *= self.TOTAL_TO_UNION_RATIO_ADJUSTMENT
            is_matched = total > union

            # MATCH CRITERIA
            logger.info('----------- GROUP MATCH -----------')
            logger.info(f'Total: {total}')
            logger.info(f'Union: {union}')
            logger.info(f'Closest Meeting: {closest_meeting}')
            logger.info(f'Current Meeting: {meeting}')
            logger.info(f'Is Matched: {is_matched}')
            if is_matched:
                return group
        return []

    def get_serializable_data(self):
        """
        Pass this data back into deserialize below to re-instantiate the
        object.
        """
        if not self.is_processed:
            raise Exception('Unprocessed MeetingSet cannot be serialized.')
        serialized_groups = []
        for group in self.groups:
            serialized_meetings = []
            for meeting in group:
                meeting_dict = {
                    'unidentifiable': meeting.unidentifiable,
                    'attendees': [(s.name, s.zoom_attendance_report) for s in meeting.attendees],
                    'datetime': meeting.datetime.isoformat(),
                    'topic': meeting.topic,
                    'search_confidence': meeting.SEARCH_CONFIDENCE_THRESHOLD
                }
                serialized_meetings.append(meeting_dict)
            serialized_groups.append(serialized_meetings)
        return serialized_groups

    def serialize(self):
        return json.dumps(self.get_serializable_data())

    @ classmethod
    def deserialize(cls, data: list):
        """
        Reconstruct the class from data previously generated by the serialize
        method. The only caveat is that the original raw data is lost, but
        that should never be a problem since the class enforces that
        self.process() be called before serialization.

        If retention of raw data across serialization and deserialization, it
        can be easily added in a subclass but for my current purposes, it would
        be a waste.

        VERY DIFFICULT BUG THAT CONFUSED ME FOR A LONG TIME:

        If `cls` is instantiated after Student.zoom_attendance_report are
        deserialized and assigned to the student class THOSE ATTENDANCE
        REPORTS ARE ERASED! This happens because `cls.__init__()` initializes
        the zoom attendance reports on all helper students, thus ERASING
        all that data.

        I've since changed the __init__ method to be a bit more gentle and
        check if a zoom_attendance_report is already there just in case, but
        that was a nasty, nasty bug.
        """
        self = cls([])
        groups = []
        all_meetings = []

        # for each group
        for serialized_group in data:
            group_meetings = []

            # for each meeting_dict
            for meeting_dict in serialized_group:
                unidentifiable = meeting_dict['unidentifiable']
                attendees = []

                # students should always match since names have been cleaned.
                for st_name, st_zar in meeting_dict['attendees']:
                    st = self.helper.find_nearest_match(st_name, auto_yes=True)
                    if not st:
                        raise Exception('unexpected no st')
                    st.zoom_attendance_report = st_zar
                    attendees.append(st)

                # reconstruct meeting object
                meeting = Meeting('')
                meeting.SEARCH_CONFIDENCE_THRESHOLD = meeting_dict['search_confidence']
                meeting.topic = meeting_dict['topic']
                meeting.datetime = datetime.datetime.fromisoformat(
                    meeting_dict['datetime']
                )
                meeting.unidentifiable = unidentifiable
                meeting.attendees = attendees


                # make appends to reconstruct MeetingSet later
                group_meetings.append(meeting)
                all_meetings.append(meeting)
            groups.append(group_meetings)

        # reconstruct MeetingSet
        self.groups = groups
        self.meetings = all_meetings
        return self

    @ staticmethod
    def deserialize_from_string(jsonstr: str):
        return MeetingSet.deserialize(json.loads(jsonstr))



class DynamicDateColorer:
    """
    Class meetings after October 4, 2020 changed from 30 to 45 minutes.
    Hence, the coloring of cells needs to be responsive.

    Thresholds can be easily changed by modifying the constants in init, and
    extensibility to allow for custom thresholds is definitely a possibility.
    """

    def __init__(self, *a, **kw):
        # define constants
        self.OCTOBER_4_2020 = 1601769600  # timestamp
        self._30_MIN_GREEN = 25
        self._30_MIN_YELLOW = 18
        self._45_MIN_GREEN = 40
        self._45_MIN_YELLOW = 30

        def fill(colorcode):
            return PatternFill(
                fill_type='solid',
                start_color=colorcode,
                end_color=colorcode
            )

        self.colors = {
            'green': fill('18fc03'),
            'yellow': fill('fcf403'),
            'red': fill('fc0303'),
        }

    def get_color(self, duration: int, timestamp: int):
        if timestamp < self.OCTOBER_4_2020:
            return self._30_min_class_color(duration)
        return self._45_min_class_color(duration)

    def _30_min_class_color(self, duration):
        if duration > self._30_MIN_GREEN:
            return self.colors.get('green')
        if duration > self._30_MIN_YELLOW:
            return self.colors.get('yellow')
        return self.colors.get('red')

    def _45_min_class_color(self, duration):
        if duration > self._45_MIN_GREEN:
            return self.colors.get('green')
        if duration > self._45_MIN_YELLOW:
            return self.colors.get('yellow')
        return self.colors.get('red')

    def get_color_avg(self, duration):
        if duration > 30:
            return self.colors.get('green')
        if duration > 15:
            return self.colors.get('yellow')
        return self.colors.get('red')


class WorkbookWriter:
    """
    Writes a MeetingSet into an excel report

    Dynamically Grouped Sheet:
        The master sheet contains a summary of all data in one place.
        Students are listed from top to bottom by dynamic grouping,
        and each meeting date gets a column to the right of the student
        name column which shows the students minutes of attendance and is

    All Students Sheet:
        All students regardless of whether they ever attended a zoom
        meeting with an ov
        Each group gets its own sheet for its attendance.

    Highlights Sheet(s):
        Gives some helpful highlights:
            - Students in the top and bottom 10th percentile of attendance
                across all groups.
            - Group with the best & worst attendance (this is only measured
                against the historical max attendance for that group, since
                the total size of the group is unknown)
            - Students whose name could not be matched. Some students use
                weird names. This program ignores those students if they
                are truly unidentifiable, but they get dumped onto the
                highlight sheet so you can at least see how many anonymous
                students there are, and what names they are using.

    Raw Data Sheet:
        Simple view of the unsorted raw data.
    """

    def __init__(self, meeting_set: MeetingSet, *a, **kw):
        self.meeting_set = meeting_set

        # init workbook
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # sheet writer classes do the heavy lifting.
        self.sheet_writer_classes = [
            MainSheetWriter,
            HomeroomSummaryWriter,
            HighlightSheetWriter,
            RawDataWriter
        ]

    def generate_report(self) -> Workbook:
        """
        Call funcs below to generate each sheet, then return the workbook.
        """
        # write sheets
        for SheetWriter in self.sheet_writer_classes:
            logger.info(f'Writing with SheetWriter class: {SheetWriter}')
            wr = SheetWriter(self.meeting_set, self.workbook.create_sheet())
            wr.write_sheet()

        return self.workbook


class BaseSheetWriter(DynamicDateColorer):
    def __init__(
            self,
            meeting_set: MeetingSet,
            sheet,
            start_row=1,
            title='Sheet'
    ):
        super().__init__()
        self.meeting_set = meeting_set
        self.sheet = sheet
        self.groups = meeting_set.groups
        self.sheet.title = title
        self.cur_row = start_row

    def write_sheet(self):
        """
        Write data into the sheet provided to init.
        """
        raise NotImplementedError

    def write_cell(self, *,
                   value: str, row=None, col: int, font=None, fill=None
                   ):
        """
        Utility for writing to a single cell with styles. Does not incremenet
        self.cur_row.

        All positional args are captured for clarity. Use keyword args only.
        """
        row = row if row else self.cur_row          # override default row value
        cell = self.sheet.cell(row=row, column=col)  # select column
        cell.value = value                          # write value
        if font:                                    # set font
            cell.font = font
        if fill:                                    # set fill
            cell.fill = fill


class MainSheetWriter(BaseSheetWriter):
    """
    Note that this class does not support dynamic coloring, it uses the
    constants defined in ExcelStyles above.
    """

    def __init__(self, *a, **kw):
        super().__init__(*a, **kw, title='Attendance by Dynamic Grouping')

        self.cur_meetings = []
        self.cur_group_students = []
        self.cur_headers = []

        # map of meeting header strings to their timestamp, for later.
        self.name_to_meeting_map = {
            m.__str__(): m for m in self.meeting_set.meetings }

    def write_sheet(self):
        """
        Orchestrate private functions below.
        """
        self._write_header()
        self.cur_row += 1
        self._explain_dynamic_groupings()
        self.cur_row += 2
        for group in self.groups:

            logger.debug(f'Main sheet writer writing group:\n{group}')

            self.cur_meetings = group

            # get all students from the group
            students = set()
            for meeting in group:
                meet_students = set(meeting.attendees)
                students.update(meet_students)
            self.cur_group_students = students

            self._write_section_title()
            self._write_group_headers()
            self._write_rows()
            self.cur_row += 2  # leave 2 blank rows between groups

    def _write_header(self):
        """
        Write title and key information; this is the header for the whole
        document.
        """
        # header information
        a1 = self.sheet['A1']
        a1.value = self.sheet.title
        a1.font = Font(size=30, bold=True, name='Cambria')
        b1 = self.sheet['A2']
        b1.value = 'Key'
        b1.font = Font(size=16, name='Calibri')

        # fill in key
        b2, c2 = self.sheet['A3':'B3'][0]
        b2.fill = self.colors['green']
        c2.value = (
            f'Green: Student attended for at least {self._30_MIN_GREEN} '
            f'minutes before 10/4/2020, or {self._45_MIN_GREEN} after '
            '10/4/2020.'
        )
        b3, c3 = self.sheet['A4':'B4'][0]
        b3.fill = self.colors['yellow']
        c3.value = (
            f'Yellow: Student attended for at least {self._30_MIN_YELLOW} '
            f'minutes before 10/4/2020, or {self._45_MIN_GREEN} after '
            '10/4/2020.'
        )
        b4, c4 = self.sheet['A5':'B5'][0]
        b4.fill = self.colors['red']
        c4.value = (
            f'Red: Student attended for less than {self._30_MIN_YELLOW} before '
            f' 10/4/2020, or {self._45_MIN_YELLOW} after 10/4/2020.'
        )
        self.cur_row += 5

    def _explain_dynamic_groupings(self):
        rows = [
            'Dynamic groupings are created programmatically by looking at '
            'the percentage of overlap of different zoom meeting instances.',

            'For certain people, this can be especialy useful. For example, '
            'if you regularly meet with a subset of your students, like '
            'in a small group instruction session, ',

            'that smaller subset of the main group will be dynamically put '
            'into a different grouping on this sheet. ',

            'This is useful because you can get a birds eye view of the '
            'attendance of each TYPE of group of students that you meet '
            'with, without doing all that grouping work manually.',

            '... If everything worked as intended, that is!'

        ]
        for row in rows:
            self.write_cell(
                col=1,
                value=row
            )
            self.cur_row += 1

    def _write_section_title(self):
        """
        Before the actual column headers, write a description of the section
        overall.
        """
        group_topics = ' ,'.join(
            topic_set := {t.topic for t in self.cur_meetings})

        s = 's' if len(topic_set) > 1 else ''  # plurality
        self.write_cell(
            col=1,
            value=f'Dynamic Group{s}',
            font=Font(size=32, bold=True)
        )
        self.cur_row += 1

        self.write_cell(
            col=1,
            value=(
                f'Group contains meet topic{s}: {group_topics}'
            ),
            font=Font(size=24, italic=True)
        )
        self.cur_row += 1

    def _write_group_headers(self):
        """
        For each group, write a row of headers that describe each meeting the
        group had.
        """

        # write the last header row: column labels
        logger.debug(
            'Meetings at the point of writing headers\n '
            + '\n'.join([str(i) for i in self.cur_meetings])
        )
        self.cur_meetings.sort(key=lambda m: m.datetime)
        self.cur_headers = [m.__str__() for m in self.cur_meetings]
        # logger.debug(f'Headers assigned: {self.cur_headers}')
        write_headers = ['Last Name', 'First Name'] + self.cur_headers
        for i, topic in enumerate(write_headers):
            i += 1  # 1-based index for openpyxl
            self.write_cell(
                col=i,
                value=topic,
                font=Font(bold=True)
            )
        self.cur_row += 1

    def _write_rows(self):
        """
        Write row for each student.
        """
        students = list(self.cur_group_students)
        students.sort(key=lambda s: s.last_name)
        for student in students:

            # write student name
            for i, name in enumerate([student.last_name, student.first_name]):
                self.write_cell(
                    col=i + 1,
                    value=name,
                    font=Font(size=16)
                )

            # iterate over headers to fill in row
            for i, header in enumerate(self.cur_headers):

                # select cell
                i += 3  # offset for first and last name

                # select value
                if not (mins_attended := student.zoom_attendance_report.get(header)):
                    mins_attended = 0

                self.write_cell(
                    col=i,
                    value=mins_attended,
                    fill=self.get_color(
                        mins_attended,
                        self.name_to_meeting_map[header].datetime.timestamp()
                    ),
                    font=Font(size=16)
                )
            self.cur_row += 1


class HomeroomSummaryWriter(MainSheetWriter, HelperConsumer):

    def __init__(self, *a, **kw):
        super().__init__(*a, **kw)
        self.sheet.title = 'Attendance by Homeroom'
        self.cur_homeroom = None

    def write_sheet(self):
        self._write_header()
        self.cur_row += 2

        # for each homeroom, sorted by grade and teacher name
        for homeroom in self._sorted_homerooms():

            logger.info(f'HomeroomSummaryWriter writing {homeroom.teacher}')

            self.cur_homeroom = homeroom
            self.cur_group_students = homeroom.students

            # homeroom_meeting_set will the set of meetings that students in
            # the homeroom went to
            homeroom_meeting_set = set()
            for st in homeroom.students:

                # logger.debug(f'{st.name} has been to meetings: {st.zoom_attendance_report}')

                # update cur_meetings with all the meetings this student went
                # to
                homeroom_meeting_set.update([
                    m for m in st.zoom_attendance_report  # dict keys
                ])

            # convert set to list for compatibility with parent classs
            self.cur_meetings = [
                self.name_to_meeting_map[n] for n in homeroom_meeting_set
            ]
            # logger.debug(f'Homeroom\'s meetings before writing: {self.cur_meetings}')

            self._write_homeroom_title()
            self._write_group_headers()
            self._write_rows()
            self.cur_row += 2

    def _write_homeroom_title(self):
        self.write_cell(
            col=1,
            value=(
                f'{self.cur_homeroom.teacher}, Grade '
                + str(self.cur_homeroom.grade_level)
            ),
            font=Font(size=24, bold=True)
        )
        self.cur_row += 1

    def _write_section_title(self):
        self.write_cell(
            col=1,
            value=(
                f'{self.cur_homeroom.teacher}, Grade '
                + self.cur_homeroom.grade_level
            ),
            font=Font(size=24, bold=True)
        )

    def _sorted_homerooms(self):
        """
        Return a list of  homerooms, sorted primarily
        by homeroom grade and secondarily by homeroom teacher name.
        """
        homerooms = [h for h in self.helper.homerooms.values()]
        homerooms.sort(key=lambda h: (h.grade_level, h.teacher))
        return homerooms


class HighlightSheetWriter(BaseSheetWriter):

    def __init__(self, *a, **kw):
        super().__init__(*a, **kw, title='Highlights')

    def write_sheet(self):
        self._write_sheet_header()
        self._write_missing_students()

        # Placeholder message on charts
        self.cur_row += 3
        data = [
            'I\'m thinking about adding charts, but I want to keep things '
            'simple. If you think charts would be really useful to you, '
            'send me an email at jdevries@empacad.org and I\'ll look '
            'into it further.',
            'For example, charts of average attendance '
            'over time, standard devition in attendance duration over time '
            '(basically, consistency across the class), etc.'
        ]
        for row in data:
            self.write_cell(
                col=1,
                value=row,
                font=Font(size=16, italic=True)
            )
            self.cur_row += 1

    def _write_sheet_header(self):
        self.write_cell(
            value='Highlights',
            col=1,
            font=Font(size=32, bold=True),
        )
        self.cur_row += 2

        self.write_cell(
            value='Unmatched Names',
            col=1,
            font=Font(size=24, bold=True)
        )
        self.cur_row += 1

        self.write_cell(
            value=(
                'This report is not perfect. As you very well know, many '
                'students use names in zoom calls that have no relationship '
                'to their real name. You had an opportunity to manually match '
                'some of these names back in the web interface, but even a '
                'human can\'t necessecarily reliably match these names. '
            ),
            col=1,
        )
        self.cur_row += 1

        self.write_cell(
            value=(
                'Nevertheless, the names are stored here for your reference, '
                'and so that you have a sense of how many students are missing '
                'from the rest of this report.'
            ),
            col=1,
        )
        self.cur_row += 1

    def _write_missing_students(self):
        """
        Write missing students in a block.
        """
        unidentifiable = set()
        for group in self.groups:
            for meeting in group:
                unidentifiable.update(meeting.unidentifiable)
        start_block = copy(self.cur_row)
        cur_col = 1
        n = len(unidentifiable) // 10
        for i, name in enumerate(unidentifiable):
            i += 1
            # every n rows, move one column over and fill the same row range
            # again. n is porportional to len(unidentifiable) to make a nice block.
            if not i % n:
                cur_col += 1
                self.cur_row = copy(start_block)

            self.cur_row += 1

            # write name to current cell
            self.write_cell(
                row=self.cur_row,
                col=cur_col,
                value=name
            )

        # cleanup; set cur_row to two rows after the end of the name block
        self.cur_row += start_block + n + 2


class RawDataWriter(BaseSheetWriter):
    def __init__(self, *a, **kw):
        super().__init__(*a, **kw)
        self.sheet.title = 'Raw Data'
        self.name_to_meeting = {
            m.__str__(): m for m in self.meeting_set.meetings
        }

    def write_sheet(self):
        self._write_headers()
        self._write_data()

    def _write_headers(self):
        headers = [
            'last name',
            'first name',
            'student homeroom',
            'meeting topic',
            'meeting date',
            'meeting time',
            'attendance duration',
        ]
        for i, header in enumerate(headers):
            self.write_cell(
                col=i + 1,
                value=(header.title())
            )
        self.cur_row += 1

    def _write_data(self):
        for meeting in self.meeting_set.meetings:
            for st in meeting.attendees:
                for k, v in st.zoom_attendance_report.items():
                    meeting = self.name_to_meeting[k]
                    data = [
                        st.last_name,
                        st.first_name,
                        st.homeroom,
                        meeting.topic,
                        str(meeting.datetime.date()),
                        meeting.datetime.strftime('%H %M'),
                        v
                    ]
                    for i, d in enumerate(data):
                        self.write_cell(
                            col=i + 1,
                            value=d,
                        )
                    self.cur_row += 1
