import csv
import logging
from datetime import datetime
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from ..assignment_submissions import (
    AssignmentSubmission,
    EdpuzzleSubmission,
    FlipgridSubmission
)
from ..json_parsers import assignment_participation_audit
from ..student import Student

def parse_homeroom(path, csv_file):
    """
    Path must be pathlib Path object.
    """
    teacher = path.stem[2:]
    grade_level = path.stem[1]

    # assign indicies of id and name rows to variables "id_index" and "name
    # index."
    with open(path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            for index, item in enumerate(row):
                if item.lower() == 'id':
                    id_index = index
                if item.lower() == 'name':
                    name_index = index
        if not id_index or not name_index:
            raise Exception('Appropriate headers not found.')

    # instantiate a student for every student in the csv
    students = []
    with open(path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        reader.__next__()
        for row in reader:
            inverted_name = row[1]
            flip_index = inverted_name.index(',')
            last_name = inverted_name[:flip_index]
            first_name = inverted_name[(flip_index + 2):]

            context = {
                'first_name': first_name,
                'last_name': last_name,
                'student_id': row[0],
                'email': None,
                'homeroom': teacher,
                'grade_level': grade_level,
            }

            student = Student(context)
            students.append(student)

    return teacher, grade_level, students


def parse_group(path):
    """
    path must be a pathlib Path object.
    """
    group_name = path.stem[2:]
    grade_level = path.stem[1]
    students = []
    with open(path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            nm = row[0].split(' ')
            context = {
                'first_name': nm[0],
                'last_name': nm[1],
                'student_id': None,
                'email': None,
                'homeroom': None,
                'groups': [group_name],
                'grade_level': grade_level,
            }
            students.append(Student(context))

        return group_name, grade_level, students

class EndOfSchoolYearMixin:
    def __init__(self, homerooms=None, students=None, groups=None):
        self.homerooms = homerooms
        self.students = students
        self.groups = groups

    def write_assignments_to_workbook(self, output_path):
        """
        Students must have the attribute st.assignment; a list of AssignmentSubmission
        objects, or a subclass of AssignmentSubmission.
        """
        OUT = Workbook()
        OUT.remove(OUT.active)
        for hr in self.homerooms:
            sheet = OUT.create_sheet(title=hr.teacher)
            sheet.append(['First Name', 'Last Name'])
            hr.students.sort(key=lambda e: e.last_name)
            for st in hr.students:
                sheet.append([st.first_name, st.last_name])
            class_assignments = set()
            [class_assignments.update(st.assignments) for st in hr.students]
            for index, assignment_name in enumerate(class_assignments):
                column = index + 3
                sheet.cell(1, column).value = assignment_name
                for i, st in enumerate(hr.students):
                    try:
                        assignment_status = st.assignments[assignment_name].status
                    except KeyError:
                        assignment_status = 0
                    colors = [
                        'fc0303',  # red
                        'fcf403',  # yellow
                        '0398fc',  # orange
                        '93f542',  # lightgreen
                        '18fc03',  # green
                    ]
                    cell = sheet.cell(i+2, column)
                    cell.value = assignment_status
                    cell.fill = PatternFill(
                        start_color=colors[assignment_status],
                        end_color=colors[assignment_status],
                        fill_type='solid'
                    )
            average_column = sheet.max_column + 1
            sheet.cell(1, average_column).value = 'Avg'
            for row in range(2, sheet.max_row+1):
                values = []
                for column in range(1, sheet.max_column):
                    cell = sheet.cell(row+1, column+1)
                    if column >= 3:
                        if sheet.cell(row, column).value:
                            values.append(sheet.cell(row, column).value)
                        else:
                            values.append(0)
                average = sum(values) / len(values)
                sheet.cell(row, average_column).value = average
        OUT.save(filename=output_path)

    def match_assgt_with_students(self, helper, context, flipgrid_csv_path, edpuzzle_csv_path):
        """
        This function needs a good amount of context:
        'flipgrid_assignments' and 'edpuzzle_assignments':
            Lists of tuples. Each item should be the assignment name and the
            path to the csv of data from edpuzzle or flipgrid which will ultimately
            be sent to the constructors for their corresponding AssignmentSubmission
            objects.
        'epoch_cutoff': integer -- epoch timestamp before which assignments will be ignored.
        """
        # expect
        FLIPGRID_ASSIGNMENTS = [i[0] for i in context['flipgrid_assignments']]
        EDPUZZLE_ASSIGNMENTS = [i[0] for i in context['edpuzzle_assignments']]
        EPOCH_CUTOFF_TIME = context['epoch_cutoff']
        for st in helper.students:
            st.assignments = {}
        # assign json paths as attributes of homeroom teachers
        for homeroom in Path('google_classrooms').iterdir():
            if homeroom.name.startswith('.'):
                continue
            teacher_name = homeroom.name.split('-')[0].strip()
            if 'Mohawk' in teacher_name:
                for hr in [h for h in helper.homerooms if h.grade_level == '3']:
                    hr.json = homeroom
                continue
            try:
                teacher = [
                    hr for hr in helper.homerooms if hr.teacher == teacher_name][0]
            except:
                raise Exception(
                    f'jack: Not yet implemented\n\nLocals dump:\n\n{locals()}'
                )
            teacher.json = homeroom
        # iterate through google classroom objects
        assignments = set()
        for hr_json in Path('google_classrooms').iterdir():
            with open(hr_json, 'r') as jsn:
                gc_data = json.load(jsn)
                for post in gc_data['posts']:
                    post_epoch = datetime.strptime(
                        post['creationTime'], '%Y-%m-%dT%H:%M:%S.%fZ').timestamp()
                    if post_epoch < EPOCH_CUTOFF_TIME:
                        continue
                    try:
                        post['courseWork']['submissions']
                    except KeyError:
                        try:
                            post['courseWork']
                            logging.error(
                                f'Post with coursework has no submission attached\n\t{post}["courseWork"]')
                            continue
                        except KeyError:
                            continue
                    assignment_title = post['courseWork']['title']
                    assignments.add(assignment_title)
                    # at this point, it is certain that the current post is an assignment,
                    # which includes coursework submissions.
                    for submission in post['courseWork']['submissions']:
                        try:
                            student_name = submission['student']['profile']['name']['fullName']
                            st = helper.find_nearest_match(student_name)
                            # switch based on assignment type
                            if assignment_title in FLIPGRID_ASSIGNMENTS:
                                subm = FlipgridSubmission(
                                    assignment_title, submission, st, flipgrid_csv_path)
                                st.assignments.setdefault(subm.title, subm)
                            elif assignment_title in EDPUZZLE_ASSIGNMENTS:
                                subm = EdpuzzleSubmission(
                                    assignment_title, submission, st, edpuzzle_csv_path)
                                st.assignments.setdefault(subm.title, subm)
                            else:
                                subm = AssignmentSubmission(
                                    assignment_title, submission)
                                st.assignments.setdefault(subm.title, subm)
                        except KeyError:
                            logging.debug(
                                f'Not a coursework post: {submission}')
                            continue
                        if not 'comments' in submission:
                            continue
                        for comment in submission['comments']:
                            subm.acknowledge_comment(comment)

    def find_non_participators(self, directories):
        """
        Opens every file in each of the passed directories (csv or json), and
        identfies assignments as:

            Google classroom json exports
            Flipgrid csv exports
            Edpuzzle csv exports

        Then, the function writes overall participation data to a conditionally
        formatted excel file with openpyxl.
        """

        # give each assignment a unique id.
        # maintain a dictionary of assignments, where the UUID is the key, and
        # the value is a dictionary with keys "NAME", and "DATE"

        self.assignments = {}
        for st in self.students:
            """
            this dict will have keys:

                assignment_id: str
                assignment_completed: bool
            """
            st.assignments = {}

        for direc in directories:
            for file_path in direc.iterdir():

                # check for mac bs
                if '.csv' not in file_path.name:
                    if '.json' not in file_path.name:
                        continue

                # route file
                if direc.name == 'google_classrooms':
                    assignment_participation_audit(file_path, self)
                if direc.name == 'flipgrid':
                    continue
                    nonparticipator_audit_flipgrid(file_path, self)
                if direc.name == 'edpuzzle':
                    continue
                    nonparticipator_audit_edpuzzle(file_path, self)
