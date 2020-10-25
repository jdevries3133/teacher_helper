import csv
from datetime import datetime
from pathlib import Path
import logging
import statistics
import string
import re

from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from .csv_parser import BaseCsvParser
try:
    from .manual_zoom_matches import MANUAL_FIXES
except ImportError:
    def MANUAL_FIXES(name):
        """
        For students with silly names the algorithm cannot recognize, write
        this function in the file above to make individual corrections.
        ./manual_zoom_matches is in the .gitignore because it will contain
        student names.
        """
        return name
from .helper import Helper

helper = Helper.read_cache()


class Meeting:
    """
    Single zoom meeting. Opens and reads a CSV file, as downloaded from
    zoom.

    Attrs:
        - rows
        - topic
        - datetime

    """

    def __init__(self, path: Path):
        self.path = path
        self.attendees = []
        self.datetime = None
        self.duration = None
        self.topic = None
        self.total_participants = None
        self.open_report()

    def __repr__(self):
        outstr = (
            f'<helper.zoom_attendance_report.Meeting; {self.topic} at '
            f'{self.datetime.isoformat()}>'
        )
        return outstr

    def __str__(self):
        return self.path.stem

    def __eq__(self, other):
        # input validation
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        self_identifier = self.topic + self.datetime.isoformat()
        other_identifier = other.topic + other.datetime.isoformat()
        if self_identifier == other_identifier:
            return True
        return False

    def __gt__(self, other):
        # input validation
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        return len(self.attendees) > len(other.attendees)

    def __len__(self):
        return len(self.attendees)

    def open_report(self):
        """
        Opens, reads, and parses zoom report such that IO will not need to
        be performed again.

        assigns attributes:
        self.topic
        self.grade_level
        self.datetime
        """
        with open(self.path, 'r') as csvfile:
            rows = [r for r in csv.reader(csvfile)]
        # raise an exception if row 2 is not blank.
        # tolerate the case of row[2] = ['', '', '', ...]
        exception_message = (
            'Zoom report must contain meeting information. The '
            f'Report at {self.path} does not appear to contain '
            'meeting information.'
        )
        if rows[2]:
            rw = set(rows[2])
            if len(rw) <= 1:
                if rw.pop():
                    raise Exception(exception_message)
            else:
                raise Exception(exception_message)
        # parse the CSV, having cleared the exception check
        self.duration = int(rows[1][5])
        self.topic = rows[1][1]
        time_str = rows[1][2]
        (
            month,
            day,
            year,
            hour,
            minute,
            *rest
        ) = [int(i) for i in re.split(r'/| |:', time_str) if i.isdigit()]
        if 'pm' in time_str.lower():
            hour += 12
        self.datetime = datetime(year, month, day, hour, minute)
        for st in helper.students.values():
            st.zoom_attendance_record = {}
        grade_levels_within = set()
        for row in rows[4:]:
            duration = row[2]
            st = self.match_student(row[0])
            if not st:
                continue
            grade_levels_within.add(st.grade_level)
            st.zoom_attendance_record.setdefault(
                (self.topic + ';' + self.datetime.isoformat()),
                duration
            )
            self.attendees.append(st)
        if len(grade_levels_within) <= 1:
            self.grade_level = grade_levels_within.pop()

    def match_student(self, name):
        """
        Wrapper method that ties together the helper class's find_student
        method, the try_matching_student... method in this class, and some
        custom matches for students with really weird zoom names. Here is also
        where the student's raw zoom name is sanitized (punctuation removed).
        """
        # reference manual fixes, an optional import
        name = MANUAL_FIXES(name)

        # Rough cleaning
        # some use dot to delimit first / last name
        name.replace('.', ' ')
        # remove punctuation
        for char in string.punctuation:
            name.replace(char, '')
        # popular emoticon character
        name.replace('ω', '')
        try:
            return helper.find_nearest_match(name, auto_yes=True)
        except Warning:
            pass
        st = self.try_matching_student_within_grade(name)
        if st:
            return st
        for part in [i for i in re.split(' |.', name) if i]:
            st = self.try_matching_student_within_grade(part)
            if st:
                return st

    def try_matching_student_within_grade(self, student_name):
        """
        Unlike in the helper student matching function, this class is aware of
        the grade level of the student it is trying to match. If the helper
        method returns None, this fallback method tries to identify the
        student through process of elimination within their own grade. This
        helps match more students who only provide their first name.
        """
        first_name_match = process.extract(
            student_name,
            [
                s.first_name for s in helper.students.values()
                if s.grade_level == self.grade_level
            ],
            limit=3
        )
        if first_name_match[0][1] > 90:
            # we have a potential match! Let's see if it's truly a match
            # first, re-extract the student object from all students...

            # If the best two matches are the same name, there are two or more
            # students in the grade level with that name. It is therefore
            # impossible to perform a perfect match against only the first name
            if first_name_match[1][0] == first_name_match[0][0]:
                logging.debug(
                    f'Cannot proceed with {student_name}. More than one '
                    f'student in the {self.grade_level}th grade has the first '
                    f'name {first_name_match[0][0]}.'
                )

            # If the first two matches are not the same, that means the first
            # name is unique, and we can make a match within the grade level.
            for n, s in helper.students.items():
                if n.split(' ')[0] == first_name_match[0][0]:
                    logging.debug(
                        'Successful grade level match for '
                        + helper.students[n].name
                    )
                    return helper.students[n]


class MeetingSet:
    """

    # General Usage

    A directory full of meeting reports can be carelessly tossed into
    __init__. This class dynamically groups meetings by attendees. For example,
    if you use the same meeting topic (i.e. "Health") to meet with different
    groups of homerooms throughout the week, this class will observe the union
    of the set of attendees at each meeting instance, and if there is a
    significant intersection, those meetings are grouped.

    # Resultant Data Structure

    Meeting groupings will be accessible as a nested list (self.groups).
    Each item in the list will itself be a chronologically sorted list of
    Meeting instances. 

    The presumption is that there is no reliable way to know the full name of
    a meeting. For example, the meeting topic might be, "Health," but whoose
    homeroom is it? There's no way to know from here. Hence, this class simply
    takes a large pool of meetings and groups them. Labeling can be done
    elsewhere.

    The topics of each meeting are saved as an attribute on the
    meeting; "meeting.topic". It is possible for multiple meeting instances
    with the same topic to be split into different groups if the participants
    are different enough. This is expected such as in the example case detailed
    above. However, meetings with different topics will never be grouped
    together.

    If, for whatever reason, a teacher has meet with the same class under
    different zoom meet topics, that case could probably be dealt with in a
    more flexible subclass which modifies the "match_meeting_with_group_by_union"
    method.
    """

    def __init__(self, dir_path: Path, group_map=None, trust_topics=False):
        self.dir_path = dir_path
        self.groups = []
        self.TOTAL_TO_UNION_RATIO_ADJUSTMENT = 0.8

    def process(self):
        """
        Called by __init__; produces data structure
        """
        # append all meetings to groupings in self.groups
        for meeting in self.iter_csvs():
            match = self.match_meeting_with_group_by_union(meeting)
            if match:
                match.append(meeting)
            else:
                self.groups.append([meeting])

    def match_meeting_with_group_by_union(self, meeting: Meeting):
        """
        Given a list of students, calculate the union between that list, and
        all the other lists of students in meetings previously provided.

        Return the group whose union against the provided list is less than the
        length of the lists combined, indicating that these are two instances
        of the same group of students meeting.
        """
        for group in self.groups:
            # select closest_meeting: the past meeting whose len() is closest
            # to the current meeting.
            closest_meeting = None
            smallest_difference = 100  # initialize with random large number
            for prev_meeting in group:
                diff = abs(len(prev_meeting) - len(meeting))
                if diff < smallest_difference:
                    smallest_difference = diff
                    closest_meeting = prev_meeting

            # perform set union comparison
            cm__attendees = {s.name for s in closest_meeting.attendees}
            m__attendees = {s.name for s in meeting.attendees}
            union = len(cm__attendees | m__attendees)
            total = len(cm__attendees) + len(m__attendees)
            # MATCH CRITERIA
            matched = total > union and prev_meeting.topic == meeting.topic
            if matched:
                logging.debug(matched)
                return group
        return []

    def rename_csv_files(self):
        """
        Provide verbose names to csv file.
        """
        for report_path in self.dir_path.iterdir():
            with open(report_path, 'r') as csvf:
                rows = [r for r in csv.reader(csvf)]
            topic = rows[1][1]
            start_time = rows[1][2]
            date = start_time.split(' ')[0].replace('/', '-')
            new_name = topic + ' ' + date + '.csv'
            report_path.rename(Path(report_path.parent, new_name))

    def iter_csvs(self):
        """
        Skip anything that isn't a real csv file. Calls self.open_report(),
        which loads the report into memory as attributes of self; there is no
        need to yield anything, but it yields the path to the current csv
        for convenience.
        """
        for i in self.dir_path.iterdir():
            if i.name[-4:] != '.csv':
                continue
            if i.name[0] == "~" or i.name[0] == ".":
                continue
            yield Meeting(i)


class ExcelWriter:
    """
    Writes a MeetingSet into an excel report

    Optionally, pass attendance_duration_thresholds: dict as a kwarg. It should
    look like this:

        {
            'color: str': 'minimum minutes of attendance: int',
        }

        For example:

        {
            'red': 5,
            'yellow': 20,
            'green': 30
        }

        You must provide these colors:
            - 'red'
            - 'yellow'
            - 'green'
    """

    def __init__(self, meetings: MeetingSet, *a, **kw):
        super().__init__(*a, **kw)
        self.meetings = meetings
        self.thresholds = kw.get('attendance_duration_thresholds')
        self.validate_or_generate_thresholds()
        # init workbook
        self.WB_OUT = Workbook()
        self.WB_OUT.remove(self.WB_OUT.active)
        self.styles = {
            'h1': Font(size=30, bold=True, name='Cambria'),
            'h2': Font(size=16, name='Calibri')
        }

        def fill(colorcode):
            return PatternFill(
                fill_type='solid',
                start_color=colorcode,
                end_color=colorcode
            )

        self.fills = {
            'green': fill('18fc03'),
            'yellow': fill('fcf403'),
            'red': fill('fc0303')
        }

    def validate_or_generate_thresholds(self):
        if self.thresholds:
            for color in ['red', 'yellow', 'green']:
                assert color in self.thresholds
            assert (
                self.thresholds['red']
                < self.thresholds['yellow']
                < self.thresholds['green']
            )
            for i in self.thresholds.values():
                assert isinstance(i, int)
        else:
            # assign defaults
            self.thresholds = {}
            self.thresholds['red'] = 0
            self.thresholds['yellow'] = 15
            self.thresholds['green'] = 30

    def generate_report(self, destination: Path, thresholds=None):
        """
        Master Sheet:
            The master sheet contains a summary of all data in one place.
            The attendance log for each group is laid out from top to bottom.

        Group Sheets:
            Each group gets its own sheet for its attendance.

        Highlights Sheet(s):
            Gives some helpful highlights:
                - Students in the top and bottom 10th percentile of attendance
                    across all groups.
                - Group with the best & worst attendance (this is only measured
                    against the historical max attendance for that group, since
                    the total size of the group is unknown)
                - Students whose name could not be matched. Some students use
                    weird names. This program ignores those students if they
                    are truly unidentifiable, but they get dumped onto the
                    highlight sheet so you can at least see how many anonymous
                    students there are, and what names they are using.
        """
        # master sheet
        master_sheet = self.WB_OUT.create_sheet(title="Master Sheet")
        self.write_master_sheet(master_sheet)
        self.save_workbook(destination)

    def write_master_sheet(self, master_sheet):
        master_sheet.page_setup.fitToWidth = 1
        # header information
        a1 = master_sheet['A1']
        a1.value = 'Zoom Attendance Report'
        a1.font = self.styles['h1']
        b1 = master_sheet['A2']
        b1.value = 'Key'
        b1.font = self.styles['h2']

        # fill in key
        b2, c2 = master_sheet['A3':'B3'][0]
        b2.fill = self.fills['green']
        c2.value = (
            f'Green: Student attended for at least {self.thresholds["green"]} '
            'minutes.'
        )
        b3, c3 = master_sheet['A4':'B4'][0]
        b3.fill = self.fills['yellow']
        c3.value = (
            f'Yellow: Student attended for at least {self.thresholds["yellow"]} '
            'minutes.'
        )
        b4, c4 = master_sheet['A5':'B5'][0]
        b4.fill = self.fills['red']
        c4.value = (
            f'Red: Student attended for at least {self.thresholds["red"]} '
            'minutes.'
        )

        breakpoint()
        if self.meetings.group_dict:
            for meeting_set_name, meetings in self.meetings.group_dict.items():
                self.write_group_dict_item(
                    meeting_set_name, meetings, master_sheet, 'A6')

    @staticmethod
    def write_group_dict_item(
        meeting_name: str,
        meeting: MeetingSet,
        worksheet,
        starting_cell: str
    ):
        """
        Returns the first empty row (int) after the block it wrote.
        """
        # make sure we aren't in a column past z
        try:
            assert len(starting_cell) == 2
            assert starting_cell[1:].isnumerical()
            assert ord(starting_cell[0].upper()) < 90
        except AssertionError:
            raise Exception("This function can not write to columns past Z")
        # get cell range to iterate over
        min_col = starting_cell[0].upper()
        row = int(starting_cell[1:])
        # TODO finish this method
        raise NotImplementedError('Need to finish this method')

    def save_workbook(self, path: Path):
        self.WB_OUT.save(path.resolve())
