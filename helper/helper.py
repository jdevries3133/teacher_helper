from copy import copy
import csv
from datetime import datetime
import dbm
import logging
import os
from pathlib import Path
import json
from random import randint
import re
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import shelve

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import process

from .assignment_submissions import (
    AssignmentSubmission,
    EdpuzzleSubmission,
    FlipgridSubmission
)
from .csv_parsers import (
    parse_homeroom,
    parse_group,
    nonparticipator_audit_flipgrid,
)
from .docx_utils import regex_search
from .json_parsers import assignment_participation_audit
from .group import Group
from .homeroom import Homeroom
from .templates import soundtrap as soundtrap_template
from .student import Student

MODULE_DIR = os.path.dirname(__file__)


class Helper:
    """
    Driver for the entire module! See README.md
    """
    def __init__(self, homerooms=None, students=None, groups=None):
        self.homerooms = homerooms
        self.students = students
        self.groups = groups

    def write_cache(self):
        with shelve.open('cache', 'c') as db:
            db['data'] = self
            db['date'] = datetime.now()

    def get_regex_classroom_doc(self, doc_dir, regex, yes=False, bad_link_regex=r''):
        """
        Iterate through a list of word documents downloaded from google classroom.
        The assumption is that docs will follow the naming convention where the student's
        full name is the start of the filename, followed by a dash.

        Optionally, it can also take a bad link regex which will pull out bad links,
        so that you can follow up with students who pasted partial links.

        Returns a list of tuples of the student name paired with the regex match
        string or None.
        """
        matched_links = []
        bad_links = []
        for doc in Path.resolve(Path(doc_dir)).iterdir():
            if doc.name[-4:] != 'docx':
                continue
            student_name = doc.name[:doc.name.index('-')-1]
            st = self.find_nearest_match([student_name], debug=yes)[0]
            if not st:
                while True:
                    student_name = input('Type in the name exactly to get a good match')
                    st = self.find_nearest_match([student_name], debug=yes)[0]
                    if st:
                        break
            matches, misses = regex_search(st, doc, regex, near_match_regex=bad_link_regex)
            matched_links += matches
            if misses:
                bad_links += misses
        if bad_link_regex:
            return matched_links, bad_links
        return matched_links

    def match_assgt_with_students(self, context):
        """
        This function needs a good amount of context:
        'flipgrid_assignments' and 'edpuzzle_assignments':
            Lists of tuples. Each item should be the assignment name and the
            path to the csv of data from edpuzzle or flipgrid which will ultimately
            be sent to the constructors for their corresponding AssignmentSubmission
            objects.
        'epoch_cutoff': integer -- epoch timestamp before which assignments will be ignored.
        """
        # expect 
        FLIPGRID_ASSIGNMENTS = [i[0] for i in context['flipgrid_assignments']]
        EDPUZZLE_ASSIGNMENTS = [i[0] for i in context['edpuzzle_assignments']]
        EPOCH_CUTOFF_TIME = context['epoch_cutoff']
        for st in helper.students:
            st.assignments = {}
        # assign json paths as attributes of homeroom teachers
        for homeroom in Path('google_classrooms').iterdir():
            if homeroom.name.startswith('.'):
                continue
            teacher_name = homeroom.name.split('-')[0].strip()
            if 'Mohawk' in teacher_name:
                for hr in [h for h in helper.homerooms if h.grade_level == '3']:
                    hr.json = homeroom
                continue
            try:
                teacher = [hr for hr in helper.homerooms if hr.teacher == teacher_name][0]
            except:
                breakpoint()
            teacher.json = homeroom
        # iterate through google classroom objects
        assignments = set()
        for hr_json in Path('google_classrooms').iterdir():
            with open(hr_json, 'r') as jsn:
                gc_data = json.load(jsn)
                for post in gc_data['posts']:
                    post_epoch = datetime.strptime(post['creationTime'], '%Y-%m-%dT%H:%M:%S.%fZ').timestamp()
                    if post_epoch < EPOCH_CUTOFF_TIME:
                        continue
                    try:
                        post['courseWork']['submissions']
                    except KeyError:
                        try:
                            post['courseWork']
                            logging.error(f'Post with coursework has no submission attached\n\t{post}["courseWork"]')
                            continue
                        except KeyError:
                            continue
                    assignment_title = post['courseWork']['title']
                    assignments.add(assignment_title)
                    # at this point, it is certain that the current post is an assignment,
                    # which includes coursework submissions.
                    for submission in post['courseWork']['submissions']:
                        try:
                            student_name = submission['student']['profile']['name']['fullName']
                            st = helper.find_nearest_match([student_name], debug=True)[0]
                            # switch based on assignment type
                            if assignment_title in FLIPGRID_ASSIGNMENTS:
                                subm = FlipgridSubmission(assignment_title, submission, st)
                                st.assignments.setdefault(subm.title, subm)
                            elif assignment_title in EDPUZZLE_ASSIGNMENTS:
                                subm = EdpuzzleSubmission(assignment_title, submission, st)
                                st.assignments.setdefault(subm.title, subm)
                            else:
                                subm = AssignmentSubmission(assignment_title, submission)
                                st.assignments.setdefault(subm.title, subm)
                        except KeyError:
                            logging.debug(f'Not a coursework post: {submission}')
                            continue
                        if not 'comments' in submission:
                            continue
                        for comment in submission['comments']:
                            subm.acknowledge_comment(comment)

    def write_assignments_to_workbook(self, output_path):
        """
        Students must have the attribute st.assignment; a list of AssignmentSubmission
        objects, or a subclass of AssignmentSubmission.
        """
        OUT = Workbook()
        OUT.remove(OUT.active)
        for hr in self.homerooms:
            sheet = OUT.create_sheet(title=hr.teacher)
            sheet.append(['First Name', 'Last Name'])
            hr.students.sort(key=lambda e: e.last_name)
            for st in hr.students:
                sheet.append([st.first_name, st.last_name])
            class_assignments = set()
            [class_assignments.update(st.assignments) for st in hr.students]
            for index, assignment_name in enumerate(class_assignments):
                column = index + 3
                sheet.cell(1, column).value = assignment_name
                for i, st in enumerate(hr.students):
                    try:
                        assignment_status = st.assignments[assignment_name].status
                    except KeyError:
                        assignment_status = 0
                    colors = [
                        'fc0303',  # red
                        'fcf403',  # yellow
                        '0398fc',  # orange
                        '93f542',  # lightgreen
                        '18fc03',  # green
                    ]
                    cell = sheet.cell(i+2, column)
                    cell.value = assignment_status
                    cell.fill = PatternFill(
                        start_color=colors[assignment_status],
                        end_color=colors[assignment_status],
                        fill_type='solid'
                    )
            average_column = sheet.max_column + 1
            sheet.cell(1, average_column).value = 'Avg'
            for row in range(2, sheet.max_row+1):
                values = []
                for column in range(1, sheet.max_column):
                    cell = sheet.cell(row+1, column+1)
                    if column >= 3:
                        if sheet.cell(row, column).value:
                            values.append(sheet.cell(row, column).value)
                        else:
                            values.append(0)
                average = sum(values) / len(values)
                sheet.cell(row, average_column).value = average
        OUT.save(filename=output_path)

    def find_nearest_match(self, student_names, debug=False):
        """
        Takes a list of student names, and returns a list of student objects
        from self.students. If there is no exact name match, it will perform a 
        fuzzy match and ask the user to resolve the ambiguity in the command line.

        If altTab is true, it will assume this is part of a gui script, and 
        "command-tab" the user in and out of the input 
        """
        for index, name in enumerate(copy(student_names)):
            # direct match(es)
            matches = [(s.name, s) for s in self.students if s.name == name]
            if matches:
                # single direct match
                if len(matches) == 1:
                    student_names[index] = matches[0][1]
                # multiple matches
                else:
                    for match in matches:
                        done = False
                        if not debug:
                            print('-' * 80)
                            print('\nDo these names match? (y/n/p) (yes, no, pass)\n')
                            print(name + '\t' + match[0] + '\n')
                        while True:
                            if debug:
                                yn = 'y'
                            else:
                                yn = input().lower()
                            if yn == 'y':
                                student_names[index] = match[1]
                                break
                                done = True
                            elif yn == 'n':
                                break
                            elif yn == 'p':
                                pass
                            else:
                                print('Please enter y, n, or p')
                        if done:
                            break
            # no match
            else:
                qset = process.extractOne(name, [s.name for s in self.students])
                if not debug:
                    print('-' * 80)
                    print('\nDo these names match? (y/n/p) (yes, no, pass)\n')
                while True:
                    if debug:
                        u_in = 'y'
                    else:
                        u_in = input(name + '\t' + qset[0] + '\n').lower()
                    if u_in == 'y':
                        student_names[index] = [s for s in self.students if s.name == qset[0]][0]
                        break
                    elif u_in == 'n':
                        break
                    elif u_in == 'p':
                        break
                    else:
                        print('Please enter "y" or "n."')
        for name in copy(student_names):
            if not isinstance(name, Student):
                print(f'{name} was deleted because they had no match.')
                student_names.remove(name)
        return student_names  # now converted to Student class instances

    def resolve_missing_st_ids(self):
        """
        Iterate through students who have no student ID. Perform a fuzzy match
        on the pool of students, and ask the user if it is a match. No-ID case
        often arises if a group list is imported with nicknames. Often, student
        ID's are critical for LMS endpoints.

        For now, if the user says "n," the student is just deleted. Obviously
        the assumption is that students who are not part of the school are not
        in any groups. This function must be modified to handle students who are group members that are not students in the school.
        """

        st_mis = [s for s in self.students if not s.student_id]
        good_sts = [s for s in self.students if s.student_id]

        for st in st_mis:

            qset = process.extractOne(st.name, [s.name for s in good_sts])
            print('\nDo these names match? (y/n)\n')
            u_in = input(st.name + '\t' + qset[0] + '\n').lower()

            while True:

                if u_in == 'y':
                    st_good = [s for s in good_sts if s.name == qset[0]][0]
                    st.student_id = st_good.student_id
                    st.name = st_good.name
                    st.__dict__.setdefault('homeroom', st_good.homeroom)
                    break

                # delete the student; if they have no id-bearing match,
                # they probably moved away
                if u_in == 'n':
                    self.students.remove(st)
                    break

                # only accept "y" or "n"
                else:
                    print('Please enter "y" or "n"')

        # make sure that every student now has an ID; critical for genesis
        # API endpoints.
        if [s for s in self.students if not s.student_id]:
            raise Exception("Didn't handle all missing student id's")

        return self.students

    def find_non_participators(self, directories):
        """
        Opens every file in each of the passed directories (csv or json), and
        identfies assignments as:

            Google classroom json exports
            Flipgrid csv exports
            Edpuzzle csv exports

        Then, the function writes overall participation data to a conditionally
        formatted excel file with openpyxl.
        """

        # give each assignment a unique id.
        # maintain a dictionary of assignments, where the UUID is the key, and
        # the value is a dictionary with keys "NAME", and "DATE"

        self.assignments = {}
        for st in self.students:

            """
            this dict will have keys:

                assignment_id: str
                assignment_completed: bool
            """
            st.assignments = {}

        for direc in directories:
            for file_path in direc.iterdir():

                # check for mac bs
                if '.csv' not in file_path.name:
                    if '.json' not in file_path.name:
                        continue

                # route file
                if direc.name == 'google_classrooms':
                    assignment_participation_audit(file_path, self)
                if direc.name == 'flipgrid':
                    continue
                    nonparticipator_audit_flipgrid(file_path, self)
                if direc.name == 'edpuzzle':
                    continue
                    nonparticipator_audit_edpuzzle(file_path, self)

    def email(self, students, template_flag):
        if template_flag == 'soundtrap':

            no_pass = [s for s in students if not hasattr(s, 'soundtrap_password')]
            if no_pass:
                nl = '\n'
                raise Exception(
                    f'The following students do not have a password '
                    f'assigned to them:\n{[(i.name + nl) for i in no_pass]}'
                )

            # init ssl
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL('smtp.gmail.com', port=465, context=context) as server:
                server.login(
                    os.getenv('GMAIL_USERNAME'),
                    os.getenv('GMAIL_PASSWORD'),
                )

                for st in students:
                    me = 'john.devries@sparta.org'
                    you = st.email

                    msg = MIMEMultipart('alternative')
                    msg['Subject'] = 'Your Soundtrap Account is Ready!'
                    msg['From'] = me
                    msg['To'] = you
                    text, html = Email.soundtrap_template(st.first_name, st.email, st.soundtrap_password)

                    part1 = MIMEText(text, 'plain')
                    part2 = MIMEText(html, 'html')

                    msg.attach(part1)
                    msg.attach(part2)

                    resp = server.sendmail(me, you, msg.as_string())
                    print(resp)

        return 0
          
    def soundtrap_update(self, input_path, output_path, update_path, debug=False):

        with open(update_path, 'r', encoding='utf-8-sig') as update_in:
            reader = csv.reader(update_in)
            next(reader)
            before_update = [row for row in reader]

        with open(input_path, 'r', encoding='utf-8-sig') as csv_in:
            read_in = csv.reader(csv_in)
            next(read_in)
            names = []
            for row in read_in:
                if row[4][0] == 'D':
                    continue
                    
                student_name = row[2] + ' ' + row[3]
                names.append(student_name)

        students = self.find_nearest_match(names, debug=debug)

        new_accounts = []
        for st in students:
            st.soundtrap_password = st.last_name.lower() + st.first_name.lower() + str(randint(10, 99))
            new_row = [st.first_name, st.last_name, st.email, st.soundtrap_password]
            if new_row in before_update:
                continue
            new_accounts.append(new_row)

        for acc in before_update:
            acc.append('old')

        for acc in new_accounts:
            acc.append('new')

        to_dict = {}
        full_list = before_update + new_accounts
        for index, i in enumerate(full_list):
            key = i[0] + i[1]
            to_dict.setdefault(key, i)

        non_dup_old = [i[:-1] for i in to_dict.values() if i[4] == 'old']
        non_dup_new = [i[:-1] for i in to_dict.values() if i[4] == 'new']

        all_accounts = non_dup_old + non_dup_new
        all_accounts.sort()
        non_dup_new.sort()

        with open(output_path, 'w') as output:
            wr = csv.writer(output)
            wr.writerow(['First name', 'Last name', 'Email', 'Password'])
            wr.writerows(non_dup_new)

        with open(Path(update_path.parent, 'soundtrap_updated.csv'), 'w') as update:
            wr = csv.writer(update)
            wr.writerow(['First name', 'Last name', 'Email', 'Password'])
            wr.writerows(all_accounts)

        return students

    def from_classroom(self, regex, flag, attribute, classroom_path):
        """
        Utility function for getting assignment data or grades from google
        classrom json repository. Searches flag (assignments, posts, comments)
        for the regex, and assigns the value of the result to the student as
        whatever string is passed as atribute.

        classroom_path must be pathlib Path object
        """
        submissions = []
        for path in classroom_path.iterdir():
            with open(path, 'r') as jsn:
                obj = json.load(jsn)
            if flag == "assignment":
                for post in obj['posts']:
                    if 'courseWork' in post.keys():
                        mo = re.fullmatch(regex, post['courseWork']['title'])
                        if mo:
                            submissions += post['courseWork']['submissions']

        for submission in submissions:
            try:
                submitter = self.find_nearest_match([submission['student']['profile']['name']['fullName']])[0]
                submitter.__dict__.setdefault(attribute, submission)
                print(f'set attribute {attribute} on student {submitter.name} for submission\n{submission}\n\n')
            except IndexError:
                pass

        return 0

    @classmethod
    def new_school_year(cls, csvdir):
        """
        Pass an absolute path to a directory full of csv files for all homerooms
        and miscellaneous groups. Note that these CSV files must follow a
        fairly strict naming convention as described in this module's readme.
        """
        homerooms = []
        students = []
        groups = []
        # deal with all homerooms, pass over groups
        for filename in os.listdir(csvdir):
            # skip .DS_store and other nonsense
            if '.csv' not in filename:
                continue
            path_to_csv = Path(csvdir, filename)
            # call parsing functions
            if path_to_csv.stem[0] == 'h':
                teacher, grade_level, this_students = parse_homeroom(path_to_csv)
                for st in copy(this_students):
                    if st.student_id in [s.student_id for s in students]:
                        match = Helper.get_matching_student(st, students, homeroom.teacher)
                        this_students.remove(st)
                        this_students.append(match)
                    else:
                        students.append(st)
                homeroom = Homeroom(teacher, grade_level, this_students)
                homerooms.append(homeroom)
                print('homeroom', homeroom.teacher, 'has', str(len(homeroom.students)), 'students')
            elif path_to_csv.stem[0] == 'g':
                # deal with groups once all the homerooms are done to help
                # with the duplicate student problem
                continue
            else:
                raise Exception(f'Unable to parse file: {filename}')
        for filename in os.listdir(csvdir):
            # skip .DS_store and other nonsense
            if '.csv' not in filename:
                continue
            path_to_csv = Path(csvdir, filename)
            if path_to_csv.stem[0] == 'g':
                name, grade_level, this_students = parse_group(path_to_csv)
                for st in copy(this_students):
                    if st.name in [s.name for s in students]:
                        match = Helper.get_matching_student(st, students)
                        match.groups += st.groups
                        this_students.remove(st)
                        this_students.append(match)
                    else:
                        students.append(st)
                group = Group(name, grade_level, this_students)
                groups.append(group)
        new_helper = cls(homerooms, students, groups)
        new_helper.write_cache()
        return new_helper

    @staticmethod
    def get_matching_student(st, students, append_homeroom=None):
        """
        THE PURPOSE OF THIS FUNCTION is to assist the new_school_year function.
        Duplicate
        Look through the list of students. Find and return the already parsed
        instance of that student so that they can be linked to this new context.

        Some students are in two homerooms for some reason, so if this function
        is called in the homeroom parsing code block, the append_homeroom
        argument should be whatever homeroom is currently being parsed. That
        homeroom will be added to the student instance as st.alt_homeroom.
        """

        for st_match in students:
            if st.student_id:
                if st_match.student_id == st.student_id:
                    if append_homeroom:
                        st_match.alt_homeroom = append_homeroom
                        return st_match
                    else:
                        return st_match

            elif st.name:
                if st.first_name == st_match.first_name and st.last_name == st_match.last_name:
                    if append_homeroom:
                        st_match.alt_homeroom = append_homeroom
                        return st_match
                    else:
                        return st_match


        raise Exception(
            f'Match for {st.name} not found. Only call this function if you '
            'are sure that there is a matching student in the list passsed as '
            'the second argument.'
        )


    @staticmethod
    def read_cache(check_date=True):
        """
        This static method returns a class because I like to break the rules.
        there's a reason for the rules; this garbage doesn't work
        """
        with shelve.open(os.path.join(MODULE_DIR, 'cache'), 'r') as db: # todo: refactor so that the cache is written below the base directory of the package, so that I can use relative file paths
            data = db['data']
            date = db['date']

        if check_date and (datetime.now().month in range(9, 12) and date.month in range(1, 7)):
            raise Exception(
                'It appears that the cache is from last school year. Please\n'
                'provide new data, re-instantiate, and re-write cache.'
            )

        return data

    @staticmethod
    def cache_exists():
        try:
            shelve.open(os.path.join(MODULE_DIR, 'cache'), 'r')
            return True
        except dbm.error:
            return False